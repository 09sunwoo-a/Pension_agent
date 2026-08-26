"""타겟 고객 룰베이스 원본(xlsx) → targets.json 변환기 (멱등).

원본은 저장소 루트의 `IRP_타겟고객_룰베이스_v1.xlsx` — 기획자가 행내 원문 문서
(IRP 텐션 UP-②③④⑤)를 읽고 «누구를 사후관리 타겟으로 볼 것인가»를 타겟 14종으로
정규화해 확인해준 표다. 지금까지 요건 판정 임계값의 근거는 지식베이스 세그먼트의
`condition_text` 원문 하나뿐이었고, 원문이 말하지 않는 값(코드의 60% 같은)은 팀 결정
대기로 남아 있었다 — 이 표가 그 자리를 메운다.

**이 표가 코드보다 위다.** 요건 임계값이 이 표와 어긋나면 코드가 틀린 것이다. 다만 표가
스스로 «근거등급»으로 자기 확신도를 밝히므로, 그 등급을 지우지 않고 그대로 싣는다:

    A  문서 직접명시   — 원문에 대상과 임계값이 그대로 적혀 있다 (고유계정대 50% 이상)
    B  직접조건 통합   — A 를 상위 캠페인군으로 묶은 것 (TG-001 OR TG-003)
    C  조사결과 기반   — 이탈고객 «조사 비중»이 근거다. 개인 임계값이 아니다 (적극투자형 51%)
    D  설계 제안 포함  — 원문에 없는 임계값을 기획자가 제안했다 (원리금보장 80%, 12개월, 30일)

    Production  직접조건 기반 적용        Observation  관찰군(이탈 단정 금지)
    Pilot       제안 임계값의 시험 적용

D·Pilot 값을 A 와 같은 얼굴로 화면에 세우면 "행내 기준"이라 오해된다. 등급을 레코드에
남겨 두는 것은 그것을 구분해 표기할 수 있게 하기 위해서다.

값의 가공은 하지 않는다 — 조건식은 문자열 그대로 싣는다. 조건식을 여기서 파이썬 판정으로
번역하면 "원본이 무엇이었나"가 JSON 에서 사라지고, 번역이 틀렸을 때 표를 다시 봐도
드러나지 않는다. 판정은 `strategy_agent/customer.py::conditions()` 가 코드로 쓰고,
어떤 타겟을 근거로 삼았는지는 그 자리에 TARGET_ID 로 적는다.

실행 (src/ 에서):

    python -m scripts.import_targets

openpyxl 이 필요하다(변환할 때만 — 적재·테스트는 산출된 JSON 만 읽으므로 불필요).
"""

from __future__ import annotations

import json
from datetime import date, datetime
from typing import Any

from pension_agent import config

SOURCE_XLSX = config.REPO_ROOT / "IRP_타겟고객_룰베이스_v1.xlsx"
OUTPUT_JSON = config.TARGETS_JSON

#: 시트 → 산출 JSON 의 키. «요약» 시트는 다른 시트의 집계라 싣지 않는다(중복 원장 금지).
_SHEETS = {
    "타겟_Master": "targets",
    "선정조건_Detail": "conditions",
    "액션_Master": "actions",
    "IF_THEN_Rule": "rules",
    "코드_거버넌스": "codes",
    "원천문서": "sources",
}

#: 각 시트에서 헤더 행이 시작되는 위치를 찾는 열쇠말. 시트마다 제목·부제 2~3행이 앞에
#: 붙어 있고 그 줄 수가 시트마다 달라, 행 번호로 세면 시트가 하나만 늘어도 어긋난다.
_HEADER_KEYS = {
    "targets": "TARGET_ID",
    "conditions": "COND_ID",
    "actions": "ACTION_ID",
    "rules": "RULE_ID",
    "codes": "코드",
    "sources": "SOURCE_ID",
}


def _cell(v: Any) -> Any:
    """엑셀 셀 → JSON 값. 날짜는 ISO 문자열, 공백뿐인 문자열은 None 으로."""
    if isinstance(v, datetime):
        return v.isoformat(sep=" ")
    if isinstance(v, date):
        return v.isoformat()
    if isinstance(v, str):
        s = v.strip()
        return s or None
    return v


def _table(ws, header_key: str) -> list[dict[str, Any]]:
    """헤더 행을 열쇠말로 찾아 그 아래를 레코드로 읽는다. 전 열 공백인 행은 건너뛴다."""
    rows = [[_cell(c) for c in r] for r in ws.iter_rows(values_only=True)]
    head_at = next(i for i, r in enumerate(rows) if header_key in r)
    header = [h for h in rows[head_at] if h is not None]
    out: list[dict[str, Any]] = []
    for r in rows[head_at + 1:]:
        if not any(c is not None for c in r):
            continue
        out.append({h: r[i] for i, h in enumerate(header)})
    return out


def main() -> None:
    from openpyxl import load_workbook  # 변환할 때만 필요하다

    if not SOURCE_XLSX.is_file():
        print(f"원본이 없습니다: {SOURCE_XLSX}")
        raise SystemExit(1)

    wb = load_workbook(SOURCE_XLSX, data_only=True)
    doc: dict[str, Any] = {
        "meta": {
            "source": SOURCE_XLSX.name,
            "note": "기획자가 행내 원문(IRP 텐션 UP-②③④⑤)을 정규화해 확인해준 "
                    "사후관리 타겟 룰베이스. 요건 임계값의 상위 기준이다.",
            "grades": {
                "A": "문서 직접명시 — 원문에 대상·임계값이 그대로 있다",
                "B": "직접조건 통합 — A 를 상위 캠페인군으로 묶은 것",
                "C": "조사결과 기반 — 이탈고객 조사 비중이며 개인 임계값이 아니다",
                "D": "설계 제안 포함 — 원문에 없는 임계값을 기획자가 제안했다(Pilot)",
            },
        },
    }
    for sheet, key in _SHEETS.items():
        doc[key] = _table(wb[sheet], _HEADER_KEYS[key])

    OUTPUT_JSON.write_text(
        json.dumps(doc, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    by_grade: dict[str, int] = {}
    for t in doc["targets"]:
        g = str(t.get("근거등급") or "?")[0]
        by_grade[g] = by_grade.get(g, 0) + 1
    print(f"[import_targets] 타겟 {len(doc['targets'])}종 · 조건 {len(doc['conditions'])}행 · "
          f"액션 {len(doc['actions'])}행 · 룰 {len(doc['rules'])}행 · "
          f"원천문서 {len(doc['sources'])}건 → {OUTPUT_JSON.relative_to(config.SRC_ROOT)}")
    print(f"            근거등급 " + " · ".join(f"{g} {n}종" for g, n in sorted(by_grade.items())))


if __name__ == "__main__":
    main()
